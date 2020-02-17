import os
import glob
import csv

from openpyxl import load_workbook

from score_calculator import (
    NEOACScoreCalculator,
    PESScoreCalculator,
    IATScoreCalculator
)

result_path = 'data/Results'
csv_header = [
    'response_id',
    'initials',
    'age',
    'gender',
    'education',
    'occupation',
    'iat_score',
    'pes_score',
    'neoac_score_n',
    'neoac_score_e',
    'neoac_score_o',
    'neoac_score_a',
    'neoac_score_c',
]

def read_from_individual_excels():
    '''
    Function takes files from the data folder.
    Files that are in the form of one excel per survey
    Sends excel files through score_aggregator that
    get all three test scores.

    Prints all the data to an excel file. This function
    is a glue function
    '''
    gender_files = {
        'male': glob.glob('data/Male/*.xlsx', recursive=True),
        'female': glob.glob('data/Female/*.xlsx', recursive=True)
    }


    for gender, files in gender_files.items():
        gender_result_file = open('{}/{}.csv'.format(result_path, gender), 'w')
        gender_csv_writer = csv.writer(gender_result_file)
        gender_csv_writer.writerow(csv_header)

        for score_file in files:
            print('Processing {}'.format(score_file))

            wb = load_workbook(score_file)
            if len(wb.sheetnames) != 3:
                print('Please check file {} for number of tabs'.format(score_file))

            iat_column = list(wb['Sheet1'].columns)[0]
            iat_responses = list(map(lambda x: x.value, iat_column))

            pes_column = list(wb['Sheet2'].columns)[0]
            pes_responses = {idx + 1: value for idx, value in enumerate(list(map(lambda x: x.value, pes_column)))}

            neoac_column = list(wb['Sheet3'].columns)[0]
            neoac_responses = {idx + 1: value for idx, value in enumerate(list(map(lambda x: x.value, neoac_column)))}

            if (len(iat_column) != 20 or len(pes_column) != 9 or len(neoac_column) != 60):
                print('Please check file {} for survey scores'.format(score_file))
                next

            response_id, initials, age, education, occupation = os.path.basename(score_file).split('.')[0].split('_')[:5]

            iat_score = IATScoreCalculator(iat_responses).calculate_scores()
            pes_score = PESScoreCalculator(pes_responses).calculate_scores()
            neoac_score = NEOACScoreCalculator(neoac_responses, gender).calculate_scores()

            csv_row = [
                response_id,
                initials,
                age,
                gender,
                education,
                occupation,
                iat_score,
                pes_score,
                neoac_score['n'],
                neoac_score['e'],
                neoac_score['o'],
                neoac_score['a'],
                neoac_score['c'],
            ]

            gender_csv_writer.writerow(csv_row)

        gender_result_file.close()


def read_from_google_form_responses():
    '''
    Read from the google form responses
    '''

    print("\n\nProcessing form responses\n\n")
    conversion_map = {v: k for k, v in NEOACScoreCalculator.RESPONSE_MAP.items()}

    wb = load_workbook('data/form_responses.xlsx')
    result_file_name = 'data/Results/form_results.csv'

    result_file = open(result_file_name, 'w')
    csv_writer = csv.writer(result_file)
    csv_writer.writerow(csv_header)

    responses_sheet = wb[wb.sheetnames[0]]

    max_col = responses_sheet.max_column
    max_row = responses_sheet.max_row

    for i in range(2, max_row + 1):
        print('Processing row {}'.format(i))
        row = []
        for j in range(1, max_col + 1):
            cell_obj = responses_sheet.cell(row = i, column = j)
            row.append(cell_obj.value)

        response_id = row[0]
        initials = ''.join(list(map(lambda x: x[0:1], row[2].split(' '))))
        education = row[3]
        age = row[4]
        gender = row[5]
        occupation = row[6]

        iat_responses = row[11:31]
        pes_responses = {idx + 1: v for idx, v in enumerate(row[31:40])}
        neoac_responses = {idx + 1: v for idx, v in enumerate(list(map(lambda x: conversion_map[x], row[40:100])))}

        iat_score = IATScoreCalculator(iat_responses).calculate_scores()
        pes_score = PESScoreCalculator(pes_responses).calculate_scores()
        neoac_score = NEOACScoreCalculator(neoac_responses, gender).calculate_scores()

        csv_row = [
            response_id,
            initials,
            age,
            gender,
            education,
            occupation,
            iat_score,
            pes_score,
            neoac_score['n'],
            neoac_score['e'],
            neoac_score['o'],
            neoac_score['a'],
            neoac_score['c'],
        ]

        csv_writer.writerow(csv_row)

    result_file.close()



if __name__ == '__main__':
    if not os.path.exists(result_path):
        os.makedirs(result_path)

    read_from_individual_excels()
    read_from_google_form_responses()



## Output of csv has to be in the form
# ID, Initials, Age, Gender, N, E, O, A, C, PES, IAT
